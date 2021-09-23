import re

index_list = [i.start() for i in re.finditer('K', 'MVQRWLYSTNAKDIAVLYFMLAIFSGMAGTAMSLIIRLELAAPGSQYLHGNSQLFNVLVVGHAVLMIFFLVMPALIGGFGNYLLPLMIGATDTAFPRINNIAFWVLPMGLVCLVTSTLVESGAGTGWTVYPPLSSIQAHSGPSVDLAIFALHLTSISSLLGAINFIVTTLNMRTNGMTMHKLPLFVWSIFITAFLLLLSLPVLSAGITMLLLDRNFNTSFFEVSGGGDPILYEHLFWFFGQTVATIIMLMMYNDMHFSKCWKLLKKWITNIMSTLFKALFVKMFMSYNNQQDKMMNNTMLKKDNIKRSSETTRKMLNNSMNKKFNQWLAGLIDGDGYFGIVSKKYVSLEITVALEDEMALKEIQNKFGGSIKLRSGVKAIRYRLTNKTGMIKLINAVNGNIRNTKRLVQFNKVCILLGIDFIYPIKLTKDNSWFVGFFDADGTINYSFKNNHPQLTISVTNKYLQDVQEYKNILGGNIYFDKSQNGYYKWSIQSKDMVLNFINDYIKMNPSRTTKMNKLYLSKEFYNLKELKAYNKSSDSMQYKAWLNFENKWKNK')]

print(index_list)
import csv
import re
import openpyxl

protein = []
sequence = []
with open('train_sequence.txt') as fp:
    lines = fp.readlines()
    # header = lines[0].split('\t')
    # print(header)
    # for head in header:
    #     if head in needed_data:
    #         mice[head] = []

    for (i, line) in enumerate(lines):
        if i % 2 == 0:
            header = lines[i].split('\t')
            protein.append(header[0][1:])
            #print(protein_name)
            sequence.append(lines[i+1][0:-1])
#print(protein)
# print(len(protein))
# print(sequence)
# print(len(sequence))
pos_data = []
neg_data = []
windowSize = 15
pos_number = 0
neg_number = 0
final_data = []
# 打开excel文件,获取工作簿对象
wb = openpyxl.load_workbook('Training_data.xlsx')
# 从表单中获取单元格的内容
ws = wb.active  # 当前活跃的表单
#print(ws.cell(row=1, column=2))  # 获取第一行第二列的单元格
#print(ws.cell(row=1, column=2).value)
for i in range(3, 4):  # 获取1,3,4,7 行第二列的值

    protein_id = ws.cell(row=i, column=1).value
    print(protein_id)
    position = ws.cell(row=i, column=2).value
    print(position)
    index = protein.index(protein_id)
    seq = sequence[index]
    print(seq)
    math_seq = '1'
    # print(index)
    # print(seq)
    if ws.cell(row=i, column=3).value =='Suc':
        math_seq = '1'
    else:
        math_seq = '0'
    k = position
    j = k-1
    if k <= windowSize:
        piptide_a = seq[k - 1:k + windowSize]
        piptide_b = seq[2 * k - 1:k + windowSize]
        piptide_c = piptide_b[::-1] + seq[0:k - 1]
        piptide_data = piptide_c + piptide_a
        final_data.append(piptide_data)
        if math_seq == '1':
            pos_number = pos_number + 1
            code = [piptide_data, protein_id, k, 1]
            pos_data.append(code)
        if math_seq == '0':
            code = [piptide_data, protein_id, k, 0]
            neg_data.append(code)
            neg_number = neg_number + 1
    elif k > (len(seq) - windowSize):
        piptide_d = seq[j - windowSize:2 * j - len(seq) + 1]
        piptide_e = seq[j:] + piptide_d[::-1]
        piptide_f = seq[j - windowSize:j]
        final_data.append(piptide_f + piptide_e)
        if math_seq == '1':
            code = [piptide_f + piptide_e, protein_id, k, 1]
            pos_data.append(code)
            pos_number = pos_number + 1
        if math_seq == '0':
            code = [piptide_f + piptide_e, protein_id, k, 0]
            neg_data.append(code)
            neg_number = neg_number + 1
    else:
        final_data.append(seq[j - windowSize:j + windowSize + 1])
        if math_seq == '1':
            code = [seq[j - windowSize:j + windowSize + 1], protein_id, k, 1]
            pos_data.append(code)
            pos_number = pos_number + 1
        if math_seq == '0':
            code = [seq[j - windowSize:j + windowSize + 1], protein_id, k, 0]
            neg_data.append(code)
            neg_number = neg_number + 1

